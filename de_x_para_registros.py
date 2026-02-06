"""DE_X_PARA DE REGISTROS UTILIZANDO SIMILARIDADE SEMÂNTICA
    pip3 install pandas openpyxl sentence-transformers torch tqdm"""

import os
import time
# from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer, util
from tqdm import tqdm
import torch

# ==============================
# CONFIGURAÇÕES GERAIS
# ==============================
DOMINIO_PATH = '/Users/xxx/Projetos/xxx/de_x_para_registros/Dominio\
/DOMINIO.xlsx'
PASTA_ARQUIVOS = '/Users/xxx/Projetos/xxx/de_x_para_registros/Arquivos'
PASTA_SAIDA = '/Users/xxx/Projetos/xxx/de_x_para_registros/Saida'
PASTA_LOGS = '/Users/xxx/Projetos/xxx/de_x_para_registros/Logs'
LOG_PATH = os.path.join(PASTA_LOGS, 'log_tratamento.txt')
RELATORIO_PATH = os.path.join(PASTA_LOGS, 'RELATORIO_REGISTROS.xlsx')
THRESHOLD = 0.75  # nível mínimo de similaridade semântica

# ==============================
# FUNÇÃO DE LOG
# ==============================


def registrar_log(mensagem):
    """Grava uma linha no arquivo de log com data e hora."""
    with open(LOG_PATH, 'a', encoding='utf-8') as lg:
        lg.write(f'[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] \
                   {mensagem}\n')


# ==============================
# CARREGA TABELA DOMÍNIO
# ==============================
print('Carregando tabela domínio...')
registrar_log('Iniciando processamento...')

try:
    dominio_df = pd.read_excel(DOMINIO_PATH, header=None, usecols='A:D')
    dominio_df.columns = ['Descricao', 'Codigo_B', 'Codigo_C', 'Descricao_D']
    dominio_df = dominio_df.dropna(subset=['Descricao'])
except Exception as ex:
    registrar_log(f'Erro ao carregar domínio: {ex}')
    raise SystemExit(f'Erro ao carregar domínio: {ex}') from ex

# ==============================
# CARREGA MODELO DE LINGUAGEM
# ==============================
print('Carregando modelo de linguagem (português)...')
model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')

print('Gerando embeddings do domínio...')
dominio_embeddings = model.encode(
    dominio_df['Descricao'].tolist(),
    convert_to_tensor=True,
    show_progress_bar=True)

# ==============================
# FUNÇÃO DE SIMILARIDADE
# ==============================


def buscar_similaridade_semantica(texto, threshold=THRESHOLD):
    """Busca o registro mais similar semanticamente no domínio."""
    if not isinstance(texto, str) or texto.strip() == '':
        return None, None

    texto_embedding = model.encode(texto, convert_to_tensor=True)
    similaridades = util.cos_sim(texto_embedding, dominio_embeddings)[0]
    indice_max = torch.argmax(similaridades).item()
    score_max = similaridades[indice_max].item()

    if score_max >= threshold:
        linha = dominio_df.iloc[indice_max]
        return linha['Codigo_C'], linha['Descricao_D']
    else:
        return None, None


# ==============================
# PROCESSAMENTO DE ARQUIVOS
# ==============================
print('Iniciando tratamento dos arquivos Excel...')
arquivos = [
    f for f in os.listdir(PASTA_ARQUIVOS)
    if f.lower().endswith('.xlsx')
    and 'DOMINIO' not in f.upper()
    and not f.startswith('TRATADO_')]

if not arquivos:
    registrar_log('Nenhum arquivo Excel encontrado para tratamento.')
    print('Nenhum arquivo Excel encontrado.')
    raise SystemExit()

relatorio = []

for arquivo in arquivos:
    try:
        caminho_arquivo = os.path.join(PASTA_ARQUIVOS, arquivo)
        print(f'Tratando arquivo: {arquivo}')
        registrar_log(f'Iniciando tratamento de: {arquivo}')

        inicio_tempo = time.time()

        wb = load_workbook(caminho_arquivo)
        ws = wb.active

        ws['BL1'] = 'NOVA DESCRICAO REGISTRO'
        ws['BM1'] = 'NOVO CODIGO REGISTRO'

        total_linhas = ws.max_row - 1
        tratados = 0
        nao_localizadas = 0

        for row in tqdm(range(2, ws.max_row + 1), desc=f'Processando \
                       {arquivo}', unit='linha'):
            valor = ws[f'U{row}'].value
            cod_c, desc_d = buscar_similaridade_semantica(str(valor))

            if cod_c and desc_d:
                ws[f'BL{row}'] = cod_c
                ws[f'BM{row}'] = desc_d
                tratados += 1
            else:
                ws[f'BL{row}'] = '#ÑLOC'
                ws[f'BM{row}'] = '#ÑLOC'
                nao_localizadas += 1

        # Salva arquivo tratado
        novo_nome = f'TRATADO_{arquivo}'
        novo_caminho_arquivos = os.path.join(PASTA_ARQUIVOS, novo_nome)
        novo_caminho_saida = os.path.join(PASTA_SAIDA, novo_nome)
        wb.save(novo_caminho_arquivos)
        wb.save(novo_caminho_saida)

        tempo_total = time.time() - inicio_tempo
        tempo_seg = round(tempo_total, 2)
        tempo_min = round(tempo_total / 60, 2)

        percentual = round((tratados / total_linhas) *
                           100, 2) if total_linhas else 0

        relatorio.append({
            'Arquivo': arquivo,
            'Total Linhas': total_linhas,
            'tratados': tratados,
            'Não Localizadas (#ÑLOC)': nao_localizadas,
            '% Sucesso': f'{percentual}%',
            'Tempo (segundos)': tempo_seg,
            'Tempo (minutos)': tempo_min})

        registrar_log(
            f'{arquivo} tratado: {tratados}/{total_linhas} tratados | '
            f'Tempo: {tempo_min} min')
        print(
            f'{tratados} tratados | {nao_localizadas} não localizadas | \
                         Tempo: {tempo_min} min')

    except Exception as ex:
        registrar_log(f'Erro ao processar {arquivo}: {ex}')
        print(f'Erro ao processar {arquivo}: {ex}')

# ==============================
# GERA RELATÓRIO FINAL
# ==============================
if relatorio:
    relatorio_df = pd.DataFrame(relatorio)
    relatorio_df.to_excel(RELATORIO_PATH, index=False)
    registrar_log(f'Relatório final salvo em: {RELATORIO_PATH}')
    print(f'Relatório final salvo como: {RELATORIO_PATH}')

print('Processo finalizado com sucesso!')
registrar_log('Processamento concluído com sucesso.\n')
